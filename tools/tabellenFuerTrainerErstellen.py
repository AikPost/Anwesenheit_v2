import openpyxl
from sys import argv
import json
from itertools import groupby
from collections import namedtuple
from pprint import pprint

Course = namedtuple("Course",["name","trainer","startTime","endTime","location","dayVerbatim","day"])

def parseArguments() -> tuple[str,str,bool]:
    if len(argv) > 3:
        return str(argv[1]), str(argv[2]), bool(argv[3])
    
    if len(argv) > 2:
        return str(argv[1]), str(argv[2]), False

    if len(argv) > 1:
        return str(argv[1]), None, False
    
    raise Exception("Error: Please specify the filename containing the course data.")


def getTitleFrom(course, trainerName):  
    # out = courseDayVerb[0:2]
    # out += courseStart.replace(":","")
    # nameAbbreviation = courseName
    
    # replacements = {
    #     "Standard":"Std",
    #     "Latein":"Lt",
    #     "Tanzkreis":"TK",
    #     "Studenten":"Stud.",
    #     "Tanzgruppe":"TG",
    #     "Training":"Tr",
    #     " und ":"+"
    # }
    
    # for oldVal, newVal in replacements.items():
    #     nameAbbreviation = nameAbbreviation.replace(oldVal,newVal)
        
    # out += nameAbbreviation
    # out.replace(" ","_")

    courseStart=course["startTime"]
    courseDayVerb = course["dayVerbatim"]
    courseAbbr = course["abbreviation"]
    return (courseDayVerb[0:2] + " " + courseStart.replace(":","") +" " + courseAbbr)[0:31]



def main():
    filename, outputFilename, exportMasterFile = parseArguments()
    print(filename, outputFilename)
    with open(filename, "r", encoding="UTF-8") as fileReader:
        data = json.load(fileReader)
    
    pprint(data)
    periodData = data["period"]
    courseData: list[dict[str:str]] = data["courses"]
    
    # sort by recipient -> day -> startTime
    courseData.sort(key=lambda course: (course["recipient"],course["day"],course["startTime"]))
    for recipient, courses in groupby(courseData, key=lambda course: course["recipient"]):
        print(f"Creating excel file for {recipient}")
        wb = openpyxl.open("__RWC_Anwesenheitslisten_TEMPLATE_PROGRAM.xlsx")
        
        for course in courses:
            courseSheet = wb.copy_worksheet(wb["template"])
            courseSheet.title = getTitleFrom(
                course=course, 
                trainerName=recipient
            )
            
            courseSheet.cell(row=1,column=2,value=course["name"])
            courseSheet.cell(row=2,column=2,value=course["trainer"])
            courseSheet.cell(row=3,column=2,value=course["dayVerbatim"])
            courseSheet.cell(row=4,column=2,value=course["startTime"] + " - " + course["endTime"])
            courseSheet.cell(row=5,column=2,value=course["location"])
            courseSheet.cell(row=6,column=2,value=periodData["quarterStart"])
            courseSheet.cell(row=7,column=2,value=periodData["quarterEnd"])
            courseSheet.cell(row=8,column=2,value=periodData["quarterName"])
            courseSheet.cell(row=9,column=2,value=int(course["day"]))

        wb.remove(wb["template"])
        wb.save(
            f"./export/{recipient}_{periodData['quarterName']}_Anwesenheitslisten.xlsx"
        )
    
if __name__ == "__main__":
    main()




